import openpyxl


def generate_accounts_list():
    path = '/home/kajetan/Documents/pryzmat/accounts.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    accounts_list = []

    row = 3
    col = 1
    while True:
        value = ws.cell(row=row, column=col).value
        if value:
            accounts_list.append(value)
            row += 1
        else:
            break
    return accounts_list


class Account:
    def __init__(self, name, monthly_budget, revenue_target, is_graphic):
        self.name = name
        self.monthly_budget = monthly_budget
        self.revenue_target = revenue_target
        self.is_graphic = is_graphic


def get_accounts_from_excel():
    path = '/home/kajetan/Documents/pryzmat/accounts.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    accounts_list = []

    row = 3
    col = 1

    while True:
        name = ws.cell(row=row, column=col).value
        monthly_budget = ws.cell(row=row, column=col + 1).value
        revenue_target = ws.cell(row=row, column=col + 2).value
        is_graphic = ws.cell(row=row, column=col + 3).value == 'tak'
        if name:
            accounts_list.append(Account(name, monthly_budget, revenue_target, is_graphic))
            row += 1
        else:
            break
    return accounts_list
