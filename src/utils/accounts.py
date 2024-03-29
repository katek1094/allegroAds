import openpyxl


class AdsAccount:
    VALID_PRIORITIES = (1, 2, 3)

    def __init__(self, name: str, priority: int, monthly_budget: int = None, revenue_target: int = None,
                 is_graphic: bool = None):
        self.name = name.strip()
        self.monthly_budget = monthly_budget
        self.revenue_target = revenue_target
        self.priority = self._validate_priority(priority)
        self.is_graphic = is_graphic

    def __repr__(self):
        return 'AdsAccount' + self.name

    @property
    def weekly_budget(self) -> float:
        if self.monthly_budget:
            return round(self.monthly_budget / 4.345)

    @property
    def daily_budget(self) -> float:
        if self.monthly_budget:
            return round(self.daily_budget / 30.437)

    def _validate_priority(self, priority: int) -> int:
        if priority in self.VALID_PRIORITIES:
            return priority
        else:
            print(type(priority))
            print(self.name)
            raise ValueError(f'Priority has to be one of these values: {self.VALID_PRIORITIES}')


def get_all_accounts_from_excel(*, priority: int = False, only_graphic=False):
    path = '/home/kajetan/Documents/pryzmat/accounts.xlsx'  # TODO: make it global constant
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    accounts_list = []

    row = 3
    col = 1
    while True:
        account_name = ws.cell(row=row, column=col).value
        if account_name:
            account_monthly_budget = ws.cell(row=row, column=col + 1).value
            account_revenue_target = ws.cell(row=row, column=col + 2).value
            account_priority = int(ws.cell(row=row, column=col + 4).value)
            account_is_graphic = ws.cell(row=row, column=col + 5).value == 'tak'
            accounts_list.append(
                AdsAccount(account_name, account_priority, account_monthly_budget, account_revenue_target,
                           account_is_graphic))
            row += 1
        else:
            break

    if priority:
        accounts_list = [account for account in accounts_list if account.priority == priority]
    if only_graphic:
        accounts_list = [account for account in accounts_list if account.is_graphic]

    return accounts_list
