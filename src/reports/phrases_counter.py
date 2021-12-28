from openpyxl import load_workbook, Workbook
from collections import Counter

# PHRASES COUNTER IN GRAPHIC AD REPORT


wb = load_workbook(filename='/home/kajetan/Desktop/phrases.xlsx')
ws = wb.active

row = 1

values = []
not_empty = True

while not_empty:
    val = ws.cell(row=row, column=1).value
    if val:
        values += val.split(' ')
        row += 1
    else:
        not_empty = False

counts = Counter(values).most_common()
# wb.create_sheet('stats')
ws = wb['stats']

for x in enumerate(counts):
    ws.cell(row=x[0] + 1, column=1).value = x[1][0]
    ws.cell(row=x[0] + 1, column=2).value = x[1][1]

wb.save(filename='/home/kajetan/Desktop/phrases.xlsx')
