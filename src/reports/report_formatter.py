import glob
from os import stat, rename

from openpyxl import load_workbook

# MONTHLY REPORT FORMATTER

reports_path = '/home/kajetan/Documents/pryzmat/reports'
# filenames = listdir(reports_path)
from src.utils import get_all_accounts_from_excel


# for filename in filenames:
#     file_path = reports_path + '/' + filename
#     wb = load_workbook(filename=file_path)
#     stats_sheet_name = 'Statystyk_2021-11-26_2021-12-25'  # TODO: make it reactive for current date
#     del wb[stats_sheet_name]
#     ws = wb.active
#     ws.delete_cols(1, 2)
#     wb.save(file_path)

def format():
    files = glob.glob('/home/kajetan/Documents/pryzmat/reports/*')
    sorted_by_mtime_ascending = sorted(files, key=lambda t: stat(t).st_mtime)

    accounts = get_all_accounts_from_excel()

    for account, file in zip(accounts, sorted_by_mtime_ascending):
        print(file)
        print(account)
        new_file = '/home/kajetan/Documents/pryzmat/reports/' + 'raport luty ' + account.name + '.xlsx'
        rename(file, new_file)
        wb = load_workbook(filename=new_file)
        # del wb['Statystyk_2022-01-26_2022-02-25']
        ws = wb['Statystyk_2022-01-26_2022-02-25']
        ws.delete_cols(1, 2)
        ws = wb['Sprzedaż_2022-01-26_2022-02-25']
        ws.delete_cols(1, 2)
        wb.save(new_file)


format()