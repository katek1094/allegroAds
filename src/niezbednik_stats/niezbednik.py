import os
import requests
import datetime
from collections import namedtuple

from dotenv import load_dotenv
from bs4 import BeautifulSoup
from openpyxl import Workbook


def create_session():
    return requests.Session()


def login(s):
    load_dotenv()

    password = os.getenv('NIEZBEDNIK_PASSWORD')

    s.headers.update({'referer': 'https://niezbedniksprzedawcy.pl/accounts/login/'})

    soup = BeautifulSoup(s.get('https://niezbedniksprzedawcy.pl/accounts/login/').content, 'html5lib')
    csrftoken = soup.find('input', dict(name='csrfmiddlewaretoken'))['value']

    r = s.post('https://niezbedniksprzedawcy.pl/accounts/login/', {
        'csrfmiddlewaretoken': csrftoken,
        'login': 'info@pryzmat.media',
        'password': password
    })

    if r.status_code != 200:
        raise Exception('CAN NOT LOGIN INTO NIEZBEDNIK SPRZEDAWCY')


def get_account_raw_data(s: requests.Session, account_name):
    api_url = 'https://niezbedniksprzedawcy.pl/StatystykiAllegro/get_offers_stats?format=json'
    """
    params
    q - account_name
    p - page number
    l - 
    f - 1M, 2M, 3M, 1y
    offerPositionLimit - 
    """
    #     'https://niezbedniksprzedawcy.pl/StatystykiAllegro/get_offers_stats?format=json&q=handlowiec-rs-pl&p=1&l=10&f=1M&offerPositionLimit=600')
    r = s.get(api_url, params={'q': account_name, 'p': 1, 'l': 600, 'f': '1M', 'offerPositionLimit': 600})
    return r.json()


Record = namedtuple('Record', 'date value')
OfferData = namedtuple('OfferData', 'id_number records')


def format_account_data(data):
    formatted_data = []
    for offer in data['series']:
        records = []
        for record in offer['data']:
            timestamp_in_ms, position = record
            record_date = datetime.datetime.fromtimestamp(timestamp_in_ms / 1000).date()
            records.append(Record(record_date, position))
        formatted_data.append(OfferData(offer['name'], tuple(records)))

    return formatted_data


def generate_columns_for_dates_dict():
    result = {}
    start_date = datetime.date.today()
    day_count = 30
    for index, date in enumerate(start_date - datetime.timedelta(day_count - n) for n in range(day_count)):
        result[date] = index + 2
    return result


columns_for_dates = generate_columns_for_dates_dict()


def create_excel(data: [OfferData]):
    wb = Workbook()
    ws = wb.active

    # WRITE DATES
    for date, column in columns_for_dates.items():
        ws.cell(row=1, column=column).value = date.strftime('%d-%m')

    # WRITE DATA
    row = 2
    for offer_data in data:
        ws.cell(row=row, column=1).value = offer_data.id_number
        for record in offer_data.records:
            ws.cell(row=row, column=columns_for_dates[record.date]).value = record.value
        row += 1

    wb.save('/home/kajetan/Desktop/handlowiec-rs-pl.xlsx')


s = create_session()
login(s)
raw_data = get_account_raw_data(s, 'handlowiec-rs-pl')
results = format_account_data(raw_data)
create_excel(results)

# print(datetime.datetime.fromtimestamp(1639004400).strftime("%A, %B %d, %Y %I:%M:%S"))
