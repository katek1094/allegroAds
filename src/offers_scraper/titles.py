from openpyxl import load_workbook, Workbook
from collections import Counter

wb = load_workbook(filename='/home/kajetan/Desktop/titles.xlsx')
ws = wb.active

row = 1

titles = []
phrases = []
not_empty = True

while not_empty:
    title = ws.cell(row=row, column=1).value
    titles.append(title)
    if title:
        phrases += title.split(' ')
        row += 1
    else:
        not_empty = False

print(len(titles))
counts = Counter(phrases).most_common()
print(counts)
print(len(counts))


