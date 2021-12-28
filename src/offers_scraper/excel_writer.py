from openpyxl import Workbook
from openpyxl.styles.fills import PatternFill

import statistics
import datetime
import os


class ExcelWriter:
    fills = [PatternFill(patternType='solid', start_color='8c8c8c', end_color='8c8c8c'),
             PatternFill(patternType='solid', start_color='a0a0a0', end_color='a0a0a0'),
             PatternFill(patternType='solid', start_color='b4b4b4', end_color='b4b4b4'),
             PatternFill(patternType='solid', start_color='c8c8c8', end_color='c8c8c8'),
             PatternFill(patternType='solid', start_color='dcdcdc', end_color='dcdcdc'),
             PatternFill(patternType='solid', start_color='f0f0f0', end_color='f0f0f0')]
    row = 1
    col = 1
    max_level = 0

    def __init__(self, username, categories):
        print(f'excel writer started on account {username}')
        self.username = username
        self.categories = categories
        self.max_level = categories['max_level']
        self.wb = Workbook()
        self.ws = self.wb.active

        self.write_raw_data_sheet()
        self.write_prices_sheet()
        self.write_offers_sheet()
        # self.playground()

        if not os.path.exists(f'/home/kajetan/Documents/pryzmat/scraped_data/{self.username}'):
            os.makedirs(f'/home/kajetan/Documents/pryzmat/scraped_data/{self.username}')
        today = datetime.date.today().strftime('%d-%m-%Y')
        self.wb.save(f'/home/kajetan/Documents/pryzmat/scraped_data/{self.username}/{self.username}-{today}.xlsx')
        print(f'excel writer finished on account {username}')

    def playground(self):
        print(len(self.categories['offers']))  # nie bierze pod uwagę tego, że jedno id może być w kilku kategoriach
        sum = 0
        for cat in self.categories['subs']:
            sum += cat['amount']
        print(sum)  # amount bierze pod uwagę rozmiary
        ids = set()
        for offer in self.categories['offers']:
            ids.add(offer['id'])
        print(len(ids))  # rzeczywista liczba ofert - napisana pod nazwą sklepu

    def write_offers_sheet(self):
        self.wb.create_sheet('offers')
        self.ws = self.wb['offers']
        self.row = 1
        self.col = 1
        for offer in self.categories['offers']:
            self.ws.cell(row=self.row, column=self.col, value=offer['title'])
            self.ws.cell(row=self.row, column=self.col + 1, value=offer['price'])
            self.ws.cell(row=self.row, column=self.col + 2, value=offer['link'])
            self.ws.cell(row=self.row, column=self.col + 3, value=offer['id'])
            self.row += 1

    def write_prices_sheet(self):
        self.wb.create_sheet('prices')
        self.ws = self.wb['prices']
        self.row = 1
        self.col = 1
        self.ws.cell(row=self.row, column=self.col, value='wszystkie oferty')
        self.row += 1
        for price in sorted(self.generate_prices_from_offers(self.categories['offers'])):
            self.ws.cell(row=self.row, column=self.col, value=price)
            self.row += 1
        self.row = 1
        self.col += 1
        self.write_prices(self.categories['subs'], 0)

    def write_prices(self, categories, level):
        if level >= len(self.fills):
            fill_index = len(self.fills) - 1
        else:
            fill_index = level
        for sub in categories:
            # if sub['amount'] == len(sub['offers']):
            #     print(sub['name'])
            #     print(level)
            # else:
            #     print('NO')
            #     print(level)
            self.ws.cell(row=self.row, column=self.col, value=sub['name']).fill = self.fills[fill_index]
            self.row += 1
            for price in sorted(self.generate_prices_from_offers(sub['offers'])):
                self.ws.cell(row=self.row, column=self.col, value=price).fill = self.fills[fill_index]
                self.row += 1
            self.col += 1
            self.row = 1
            if len(sub['subs']):
                self.write_prices(sub['subs'], level + 1)

    @staticmethod
    def generate_prices_from_offers(offers):
        prices = []
        for offer in offers:
            prices.append(offer['price'])
        return prices

    def write_raw_data_sheet(self):
        def write_basic_info():
            self.ws.cell(column=1, row=1, value='użytkownik: ')
            self.ws.cell(column=2, row=1, value=self.username)
            self.ws.cell(column=1, row=2, value='wygenerowano: ')
            self.ws.cell(column=2, row=2, value=datetime.datetime.now())

        def write_data_labels():
            self.ws.cell(column=self.max_level + 2, row=self.row - 2, value='avg price')
            self.ws.cell(column=self.max_level + 3, row=self.row - 2, value='median price')
            self.ws.cell(column=self.max_level + 4, row=self.row - 2, value='min price')
            self.ws.cell(column=self.max_level + 5, row=self.row - 2, value='max price')

        self.ws.title = 'raw data'
        self.row = 5
        self.col = 1
        write_basic_info()
        write_data_labels()
        # self.ws.cell(column=self.max_level + 1, row=self.row - 1, value=.)
        self.write_prices_stats(self.generate_prices_from_offers(self.categories['offers']), self.row - 1)
        self.write_categories(self.categories['subs'], 0)

    def write_categories(self, categories, level):
        def add_data(cat):
            self.ws.cell(column=self.col, row=self.row, value=cat['name'])
            self.ws.cell(column=self.col + 1, row=self.row, value=cat['amount'])

        if level >= len(self.fills):
            fill_index = len(self.fills) - 1
        else:
            fill_index = level
        for category in categories:
            add_data(category)
            self.col += 1
            for x in range(self.col - 1, self.max_level + 2):
                self.ws.cell(column=x, row=self.row).fill = self.fills[fill_index]
            self.write_prices_stats(self.generate_prices_from_offers(category['offers']), self.row)
            self.row += 1
            if len(category['subs']):
                self.write_categories(category['subs'], level + 1)
            self.col -= 1

    def write_prices_stats(self, prices, row):
        self.ws.cell(column=self.max_level + 2, row=row, value=round(statistics.mean(prices), 2))
        self.ws.cell(column=self.max_level + 3, row=row, value=statistics.median(prices))
        self.ws.cell(column=self.max_level + 4, row=row, value=min(prices))
        self.ws.cell(column=self.max_level + 5, row=row, value=max(prices))
